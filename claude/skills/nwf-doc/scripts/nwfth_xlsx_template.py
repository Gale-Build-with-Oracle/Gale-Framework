"""
NWFTH XLSX Template Helper
Creates Excel spreadsheets with NWFTH branding (logo header, footer attribution)
"""

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

LOGO_PATH = os.path.join(os.path.dirname(__file__), '..', 'NWFLogo.jpg')

# NWFTH Brand Colors
NWFTH_BROWN = '3A2920'
NWFTH_LIGHT_BROWN = '8B6A53'
NWFTH_GOLD = 'E0AA2F'


def create_nwfth_workbook(sheet_name="Sheet1"):
    """
    Create a new workbook with NWFTH branding

    Args:
        sheet_name: Name for the first sheet

    Returns:
        tuple: (workbook, worksheet)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Add logo to header area
    try:
        img = XLImage(LOGO_PATH)
        img.width = 150
        img.height = 75
        ws.add_image(img, 'A1')
    except Exception:
        pass  # Logo not found, skip

    # Set footer with attribution
    ws.oddFooter.center.text = "Newly Weds Foods (Thailand)"
    ws.oddFooter.center.size = 10
    ws.oddFooter.center.font = "Arial"

    return wb, ws


def apply_header_style(cell):
    """Apply NWFTH header style to a cell"""
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color=NWFTH_BROWN, end_color=NWFTH_BROWN, fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')


def apply_data_style(cell, bold=False, color=None, wrap=True):
    """Apply NWFTH data style to a cell"""
    cell.font = Font(bold=bold, color=color or "000000", size=10)
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=wrap)


def parse_markdown_table(text):
    """
    Parse an ASCII-pipe markdown table into (headers, rows).
    Skips separator rows (|---|). Returns lists of string lists.
    Use this to convert markdown table text into real spreadsheet rows.
    """
    import re
    lines = [l.strip() for l in text.strip().split('\n') if l.strip().startswith('|')]
    result = []
    for line in lines:
        if re.match(r'^\|[-: |]+\|$', line):
            continue
        cells = [c.strip() for c in line.strip('|').split('|')]
        result.append(cells)
    if not result:
        return [], []
    return result[0], result[1:]


def auto_fit_column_width(ws, col_idx, min_w=10, max_w=80):
    """
    Set column width based on max content length in that column.
    Call AFTER all data is written.
    """
    col_letter = get_column_letter(col_idx)
    max_len = 0
    for cell in ws[col_letter]:
        if cell.value:
            lines = str(cell.value).split('\n')
            line_max = max(len(l) for l in lines)
            max_len = max(max_len, line_max)
    ws.column_dimensions[col_letter].width = max(min_w, min(max_len + 2, max_w))


def create_thin_border():
    """Create thin border style"""
    return Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )


def create_nwfth_uat_template(filename, project_name="Project", test_cases=None):
    """
    Create a UAT (User Acceptance Testing) template spreadsheet

    Args:
        filename: Output filename
        project_name: Name of the project
        test_cases: Optional list of test case dicts
    """
    wb, ws = create_nwfth_workbook(f"UAT - {project_name}")

    # Title row (below logo)
    ws['A5'] = f"UAT Test Cases - {project_name}"
    ws['A5'].font = Font(bold=True, size=16, color=NWFTH_BROWN)
    ws['A5'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells('A5:H5')

    # Metadata
    ws['A7'] = "Project:"
    ws['B7'] = project_name
    ws['A8'] = "Date:"
    ws['B8'] = datetime.now().strftime('%Y-%m-%d')
    ws['A9'] = "Prepared By:"
    ws['B9'] = "ICT - NWFTH"

    for row in [7, 8, 9]:
        ws[f'A{row}'].font = Font(bold=True, size=10)

    # Headers row
    headers = ['Test ID', 'Test Scenario', 'Test Steps', 'Expected Result', 'Actual Result', 'Status', 'Tester', 'Notes']
    header_row = 11

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        apply_header_style(cell)
        cell.border = create_thin_border()

    # Add test cases if provided
    if test_cases:
        for i, tc in enumerate(test_cases, 1):
            row = header_row + i
            ws.cell(row=row, column=1, value=tc.get('id', f'TC{i:03d}'))
            ws.cell(row=row, column=2, value=tc.get('scenario', ''))
            ws.cell(row=row, column=3, value=tc.get('steps', ''))
            ws.cell(row=row, column=4, value=tc.get('expected', ''))
            ws.cell(row=row, column=5, value=tc.get('actual', ''))

            status_cell = ws.cell(row=row, column=6, value=tc.get('status', ''))
            if tc.get('status') == 'Passed':
                status_cell.font = Font(color='00AA00', bold=True)
            elif tc.get('status') == 'Failed':
                status_cell.font = Font(color='AA0000', bold=True)

            ws.cell(row=row, column=7, value=tc.get('tester', ''))
            ws.cell(row=row, column=8, value=tc.get('notes', ''))

            # Apply borders + wrap_text for readable long-text cells
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                cell.border = create_thin_border()
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            # Minimum row height so text isn't clipped
            ws.row_dimensions[row].height = max(30, 14 * (str(tc.get('steps', '')).count('\n') + 1))

    # Set column widths — generous to avoid mid-sentence truncation
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 30

    # Set row height for header
    ws.row_dimensions[header_row].height = 25

    wb.save(filename)
    return filename


def create_nwfth_test_summary(filename, data=None):
    """
    Create a test summary report spreadsheet

    Args:
        filename: Output filename
        data: Dictionary with keys:
            - project_name: str
            - test_date: str
            - test_cases: list of dicts
    """
    if data is None:
        data = {}

    wb, ws = create_nwfth_workbook("Test Summary")

    project_name = data.get('project_name', 'N/A')
    test_date = data.get('test_date', datetime.now().strftime('%Y-%m-%d'))
    test_cases = data.get('test_cases', [])

    # Title
    ws['A5'] = "Test Summary Report"
    ws['A5'].font = Font(bold=True, size=18, color=NWFTH_BROWN)
    ws.merge_cells('A5:E5')

    # Metadata
    ws['A7'] = "Project:"
    ws['B7'] = project_name
    ws['A8'] = "Date:"
    ws['B8'] = test_date
    ws['A9'] = "Prepared By:"
    ws['B9'] = "ICT - NWFTH"

    for row in [7, 8, 9]:
        ws[f'A{row}'].font = Font(bold=True, size=10)

    # Summary statistics
    ws['A11'] = "Summary Statistics"
    ws['A11'].font = Font(bold=True, size=14, color=NWFTH_LIGHT_BROWN)

    passed = sum(1 for tc in test_cases if tc.get('status') == 'Passed')
    failed = sum(1 for tc in test_cases if tc.get('status') == 'Failed')
    total = len(test_cases)

    ws['A13'] = "Total Test Cases:"
    ws['B13'] = total
    ws['A14'] = "Passed:"
    ws['B14'] = passed
    ws['B14'].font = Font(color='00AA00', bold=True)
    ws['A15'] = "Failed:"
    ws['B15'] = failed
    ws['B15'].font = Font(color='AA0000', bold=True) if failed > 0 else Font(color='00AA00', bold=True)

    # Test results table
    ws['A17'] = "Test Results"
    ws['A17'].font = Font(bold=True, size=14, color=NWFTH_LIGHT_BROWN)

    headers = ['Test ID', 'Test Name', 'Status', 'Notes']
    header_row = 19

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        apply_header_style(cell)
        cell.border = create_thin_border()

    # Add test cases
    for i, tc in enumerate(test_cases, 1):
        row = header_row + i
        ws.cell(row=row, column=1, value=tc.get('id', ''))
        ws.cell(row=row, column=2, value=tc.get('name', ''))

        status_cell = ws.cell(row=row, column=3, value=tc.get('status', ''))
        if tc.get('status') == 'Passed':
            status_cell.font = Font(color='00AA00', bold=True)
        elif tc.get('status') == 'Failed':
            status_cell.font = Font(color='AA0000', bold=True)

        ws.cell(row=row, column=4, value=tc.get('notes', ''))

        for col in range(1, 5):
            ws.cell(row=row, column=col).border = create_thin_border()

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 35

    wb.save(filename)
    return filename


def create_nwfth_db_report(filename, transactions=None, title="Database Transaction Report"):
    """
    Create a database transaction report spreadsheet

    Args:
        filename: Output filename
        transactions: List of transaction dicts
        title: Report title
    """
    if transactions is None:
        transactions = []

    wb, ws = create_nwfth_workbook("DB Report")

    # Title
    ws['A5'] = title
    ws['A5'].font = Font(bold=True, size=18, color=NWFTH_BROWN)
    ws.merge_cells('A5:F5')

    # Metadata
    ws['A7'] = "Generated:"
    ws['B7'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws['A8'] = "Prepared By:"
    ws['B8'] = "ICT - NWFTH"

    for row in [7, 8]:
        ws[f'A{row}'].font = Font(bold=True, size=10)

    # Transaction table
    headers = ['Transaction ID', 'Table', 'Action', 'Status', 'Timestamp', 'Details']
    header_row = 10

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        apply_header_style(cell)
        cell.border = create_thin_border()

    # Add transactions
    for i, tx in enumerate(transactions, 1):
        row = header_row + i
        ws.cell(row=row, column=1, value=tx.get('id', ''))
        ws.cell(row=row, column=2, value=tx.get('table', ''))
        ws.cell(row=row, column=3, value=tx.get('action', ''))

        status_cell = ws.cell(row=row, column=4, value=tx.get('status', ''))
        if tx.get('status') == 'Success':
            status_cell.font = Font(color='00AA00', bold=True)
        elif tx.get('status') == 'Failed':
            status_cell.font = Font(color='AA0000', bold=True)

        ws.cell(row=row, column=5, value=tx.get('timestamp', ''))
        ws.cell(row=row, column=6, value=tx.get('details', ''))

        for col in range(1, 7):
            ws.cell(row=row, column=col).border = create_thin_border()

    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 35

    wb.save(filename)
    return filename


def create_nwfth_data_table(filename, sheet_name="Data", headers=None, rows=None, title=None):
    """
    Create a generic data table spreadsheet with NWFTH branding

    Args:
        filename: Output filename
        sheet_name: Sheet name
        headers: List of column headers
        rows: List of row data (list of lists)
        title: Optional title
    """
    wb, ws = create_nwfth_workbook(sheet_name)

    start_row = 5

    if title:
        ws[f'A{start_row}'] = title
        ws[f'A{start_row}'].font = Font(bold=True, size=16, color=NWFTH_BROWN)
        start_row += 2

    if headers:
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            apply_header_style(cell)
            cell.border = create_thin_border()

    if rows:
        for i, row_data in enumerate(rows, 1):
            row = start_row + i
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = create_thin_border()
                cell.alignment = Alignment(horizontal='left', vertical='center')

    # Auto-adjust column widths
    if headers:
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

    wb.save(filename)
    return filename


if __name__ == '__main__':
    import sys

    if len(sys.argv) < 2:
        print("Usage: python nwfth_xlsx_template.py [uat|test-summary|db-report|data-table]")
        sys.exit(1)

    command = sys.argv[1]

    if command == 'uat':
        create_nwfth_uat_template('uat-template.xlsx', 'Sample Project', [
            {
                'id': 'TC001',
                'scenario': 'User Login',
                'steps': '1. Enter username\n2. Enter password\n3. Click Login',
                'expected': 'User logged in successfully',
                'actual': '',
                'status': '',
                'tester': '',
                'notes': ''
            }
        ])
        print("Created uat-template.xlsx")

    elif command == 'test-summary':
        create_nwfth_test_summary('test-summary.xlsx', {
            'project_name': 'BME System',
            'test_cases': [
                {'id': 'TC001', 'name': 'Login Test', 'status': 'Passed'},
                {'id': 'TC002', 'name': 'Logout Test', 'status': 'Passed'},
                {'id': 'TC003', 'name': 'Data Export', 'status': 'Failed', 'notes': 'Timeout issue'}
            ]
        })
        print("Created test-summary.xlsx")

    elif command == 'db-report':
        create_nwfth_db_report('db-report.xlsx', [
            {'id': 'TX001', 'table': 'Orders', 'action': 'INSERT', 'status': 'Success', 'timestamp': '2026-02-10 10:00:00'},
            {'id': 'TX002', 'table': 'Customers', 'action': 'UPDATE', 'status': 'Success', 'timestamp': '2026-02-10 10:05:00'}
        ])
        print("Created db-report.xlsx")

    elif command == 'data-table':
        create_nwfth_data_table(
            'data-table.xlsx',
            headers=['ID', 'Name', 'Value', 'Status'],
            rows=[
                ['1', 'Item A', '100', 'Active'],
                ['2', 'Item B', '200', 'Inactive'],
                ['3', 'Item C', '150', 'Active']
            ],
            title='Sample Data Table'
        )
        print("Created data-table.xlsx")

    else:
        print(f"Unknown command: {command}")
        print("Usage: python nwfth_xlsx_template.py [uat|test-summary|db-report|data-table]")
